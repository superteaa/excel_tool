import tkinter as tk
from tkinter import messagebox, filedialog
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from openpyxl import load_workbook
import sys
import os

# Load the workbook and select the active worksheet
output_name = sys.argv[1]

min_v = sys.argv[2]
max_v = sys.argv[3]
sta = sys.argv[4]

wb = load_workbook(output_name)
ws = wb.active

# 从第二行开始读取数据
x_values = [float(cell.value) for cell in ws['C'][1:] if cell.value is not None]
y_values = [float(cell.value) for cell in ws['D'][1:] if cell.value is not None]


x_average = sum(x_values) / len(x_values)
y_average = sum(y_values) / len(y_values)
summ = 0.000000
mon = 0.000000

for i in range(1, len(x_values)):
    summ += (x_values[i] - x_average)*(y_values[i] - y_average)
    mon += (x_values[i] - x_average)*(x_values[i] - x_average)

k = summ/mon
b = y_average - k*x_average

output_filename = os.path.basename(output_name)

# 创建一个简单的窗口
root = tk.Tk()
if sta == "A":
    title_text = f"#{output_filename}#{min_v}-{max_v}V充电拟合曲线"
    root.title(title_text.center(100))
else:
    title_text = f"#{output_filename}#{min_v}-{max_v}V放电拟合曲线"
    root.title(title_text.center(100))

# 创建一个matplotlib图形
fig, ax = plt.subplots()
# 创建散点图
ax.scatter(x_values, y_values, s = 5)
# 添加一元一次方程的直线

x_custom = x_values  # 自定义 x 值
y_custom = [k * x + b for x in x_custom]  # 自定义 y 值，例如 y = ax + b


ax.plot(x_custom, y_custom, label='y = k*x+b', color='red')
ax.legend()  # 显示图例
# 设置横纵坐标的标识
ax.set_xlabel('CV‰')
ax.set_ylabel('SOH%')

# 将matplotlib图形嵌入到Tkinter窗口中
canvas = FigureCanvasTkAgg(fig, master=root)
canvas.draw()
canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)

# 公式
# 保留 4 位小数
k_rounded = round(k, 4)
b_rounded = round(b, 2)
formula_text = 'SOH='+str(k_rounded)+'CV'+str(b_rounded)
soh_indicates = [k * x + b for x in x_values]
differences = [(y - soh)*(y - soh) for y, soh in zip(y_values, soh_indicates)]
r_son = sum(differences)
differences = [(y - y_average)*(y - y_average) for y in y_values]
r_mon = sum(differences)
r = r_son/r_mon
r_rounded = round(r, 4)
r_text = 'R^2='+str(r_rounded)

# 复制公式的函数
def copy_formula():
    root.clipboard_clear()
    root.clipboard_append(formula_text)
    messagebox.showinfo("复制成功", "公式已复制到剪贴板")

def copy_r():
    root.clipboard_clear()
    root.clipboard_append(r_text)
    messagebox.showinfo("复制成功", "R^2值已复制到剪贴板")

# 导出图像的函数
def export_image():
    file_path = filedialog.asksaveasfilename(defaultextension=".png",
                                               filetypes=[("PNG files", "*.png"),
                                                          ("All files", "*.*")])
    if file_path:
        fig.savefig(file_path)
        messagebox.showinfo("导出成功", f"图像已保存到 {file_path}")


# 创建一个框架来包含公式和复制按钮
formula_frame = tk.Frame(root)
formula_frame.pack(side=tk.TOP, fill=tk.X, padx=190)

# 在框架中添加一个标签来展示公式，左对齐，并添加左侧填充
formula_label = tk.Label(formula_frame, text=formula_text, font=10)
formula_label.pack(side=tk.LEFT, anchor=tk.W)

# 在框架中添加一个复制按钮，右对齐，并添加水平间距
copy_button = tk.Button(formula_frame, text="复制", command=copy_formula, font=(10), height=2, width=8)
copy_button.pack(side=tk.RIGHT, anchor=tk.E, padx=50)

# 创建一个框架来包含r复制按钮
r_frame = tk.Frame(root)
r_frame.pack(side=tk.TOP, fill=tk.X, padx=190)

# 在框架中添加一个标签来展示R^2，左对齐，并添加左侧填充
formula_label = tk.Label(r_frame, text=r_text, font=10)
formula_label.pack(side=tk.LEFT, anchor=tk.W)

# 在框架中添加一个复制按钮，右对齐，并添加水平间距
copy_button = tk.Button(r_frame, text="复制", command=copy_r, font=(10), height=2, width=8)
copy_button.pack(side=tk.RIGHT, anchor=tk.E, padx=50)

# 添加一个导出按钮
export_button = tk.Button(root, text="导出图像", command=export_image, font=(10), height=2, width=8)
export_button.pack(side=tk.BOTTOM)

# 运行Tkinter事件循环
root.mainloop()

