import tkinter as tk
from tkinter import messagebox, filedialog
from openpyxl import load_workbook, Workbook
import os
import numpy as np
from multiprocessing import Pool, freeze_support

# def img_window(output_name, step_index):
#     # Load the workbook
#     print("drawing")
#     wb = load_workbook(output_name)
#     ws = wb.active
#
#     # 从第二行开始读取数据
#     x_values = [float(cell.value) for cell in ws['C'][1:] if cell.value is not None]
#     y_values = [float(cell.value) for cell in ws['D'][1:] if cell.value is not None]
#
#     # 计算平均值
#     x_average = sum(x_values) / len(x_values)
#     y_average = sum(y_values) / len(y_values)
#     summ = 0.000000
#     mon = 0.000000
#
#     # 计算斜率和b
#     for i in range(1, len(x_values)):
#         summ += (x_values[i] - x_average)*(y_values[i] - y_average)
#         mon += (x_values[i] - x_average)*(x_values[i] - x_average)
#
#     k = summ/mon
#     b = y_average - k*x_average
#
#     output_filename = os.path.basename(output_name)
#
#     # 创建一个简单的窗口
#     root = tk.Tk()
#     title_text = f"#{output_filename}"
#     root.title(title_text.center(100))
#
#     # 创建一个matplotlib图形
#     fig, ax = plt.subplots()
#     # 创建散点图
#     ax.scatter(x_values, y_values, s = 5)
#     # 添加一元一次方程的直线
#
#     x_custom = x_values  # 定义 x 值
#     y_custom = [k * x + b for x in x_custom]  # 定义 y 值
#
#     # 绘制直线
#     ax.plot(x_custom, y_custom, label='y = k*x+b', color='red')
#     ax.legend()  # 显示图例
#     # 设置横纵坐标的标识
#     ax.set_xlabel('CV‰')
#     ax.set_ylabel('SOH%')
#
#     # 将matplotlib图形嵌入到Tkinter窗口中
#     canvas = FigureCanvasTkAgg(fig, master=root)
#     canvas.draw()
#     canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
#
#     # 公式
#     # 保留 4 位小数
#     k_rounded = round(k, 4)
#     b_rounded = round(b, 2)
#     formula_text = 'SOH='+str(k_rounded)+'CV'+str(b_rounded)
#     soh_indicates = [k * x + b for x in x_values]
#     differences = [(y - soh)*(y - soh) for y, soh in zip(y_values, soh_indicates)]
#     r_son = sum(differences)
#     differences = [(y - y_average)*(y - y_average) for y in y_values]
#     r_mon = sum(differences)
#     r = 1 - (r_son/r_mon)
#     r_rounded = round(r, 4)
#     r_text = 'R^2='+str(r_rounded)
#
#     # 复制公式的函数
#     def copy_formula():
#         root.clipboard_clear()
#         root.clipboard_append(formula_text)
#         messagebox.showinfo("复制成功", "公式已复制到剪贴板")
#
#     def copy_r():
#         root.clipboard_clear()
#         root.clipboard_append(r_text)
#         messagebox.showinfo("复制成功", "R^2值已复制到剪贴板")
#
#     # 导出图像的函数
#     def export_image():
#         file_path = filedialog.asksaveasfilename(defaultextension=".png",
#                                                    filetypes=[("PNG files", "*.png"),
#                                                               ("All files", "*.*")])
#         if file_path:
#             fig.savefig(file_path)
#             messagebox.showinfo("导出成功", f"图像已保存到 {file_path}")
#
#
#     # 创建一个框架来包含公式和复制按钮
#     formula_frame = tk.Frame(root)
#     formula_frame.pack(side=tk.TOP, fill=tk.X, padx=190)
#
#     # 在框架中添加一个标签来展示公式，左对齐，并添加左侧填充
#     formula_label = tk.Label(formula_frame, text=formula_text, font=10)
#     formula_label.pack(side=tk.LEFT, anchor=tk.W)
#
#     # 在框架中添加一个复制按钮，右对齐，并添加水平间距
#     copy_button = tk.Button(formula_frame, text="复制", command=copy_formula, font=(10), height=2, width=8)
#     copy_button.pack(side=tk.RIGHT, anchor=tk.E, padx=50)
#
#     # 创建一个框架来包含r复制按钮
#     r_frame = tk.Frame(root)
#     r_frame.pack(side=tk.TOP, fill=tk.X, padx=190)
#
#     # 在框架中添加一个标签来展示R^2，左对齐，并添加左侧填充
#     formula_label = tk.Label(r_frame, text=r_text, font=10)
#     formula_label.pack(side=tk.LEFT, anchor=tk.W)
#
#     # 在框架中添加一个复制按钮，右对齐，并添加水平间距
#     copy_button = tk.Button(r_frame, text="复制", command=copy_r, font=(10), height=2, width=8)
#     copy_button.pack(side=tk.RIGHT, anchor=tk.E, padx=50)
#
#     # 添加一个导出按钮
#     export_button = tk.Button(root, text="导出图像", command=export_image, font=(10), height=2, width=8)
#     export_button.pack(side=tk.BOTTOM)
#
#     # 运行Tkinter事件循环
#     root.mainloop()

# 定义一个处理单个 check_cycle 的函数
def process_cycle(check_cycle, sheet_data, step_index, standard_sapacity):
    step_ranges = [
        (0, 0.22),
        (0.23, 0.44),
        (0.45, 0.66),
        (0.67, 0.88),
        (0.89, float('inf')),
        (0.89, float('inf')),
        (0.89, float('inf')),
        (0.89, float('inf')),
        (0.89, float('inf')),
    ]

    result = []

    if step_index != 10:
        start = step_ranges[step_index - 3][0]
        end = step_ranges[step_index - 3][1]

        step_data = [row_data['H'] for row_data in sheet_data if
                     row_data['F'] == check_cycle and
                     start <= row_data['I'] <= end and
                     row_data['G'] > 0]

        if step_data:
            average_value = np.mean(step_data)
            std_dev = np.std(step_data)
            final_cv = std_dev / average_value

            charge_capacity = next(
                (row_data['I'] for row_data in sheet_data if row_data['F'] == check_cycle and row_data['E'] == 11),
                None)
            if charge_capacity is None:
                return (check_cycle, final_cv, 0) # 如果没有找到Charge_Capacity，直接返回

            soh = charge_capacity / standard_sapacity * 100

            return (check_cycle, final_cv, soh)
        else:
            return  # 如果I列数据永远到不了，提前退出


    if step_index == 10:
        step_data = [row_data['H'] for row_data in sheet_data if
                     row_data['F'] == check_cycle and
                     row_data['G'] < 0]

        if step_data:
            average_value = np.mean(step_data)
            std_dev = np.std(step_data)
            final_cv = std_dev / average_value
            result.append((check_cycle, 10, final_cv, None))

            charge_capacity = next(
                (row_data['I'] for row_data in sheet_data if row_data['F'] == check_cycle and row_data['E'] == 11),
                None)
            if charge_capacity is None:
                return (check_cycle, final_cv, 0) # 如果没有找到Charge_Capacity，直接返回

            soh = charge_capacity / standard_sapacity * 100

            return (check_cycle, final_cv, soh)
        return



# 主程序
def main(work_name, step_index, output_name, standard_sapacity):
    # if __name__ == '__main__':
    # 加载Excel文件
    print("loading, it will take a long while, please wait. ᕙ(`▿´)ᕗ")
    workbook = load_workbook(work_name)
    sheet = workbook.active
    print("loading finished. ٩( 'ω' )و")

    # 提取工作表数据
    sheet_data = []
    for row in range(1, sheet.max_row + 1):
        row_data = {'E': sheet[f'E{row}'].value, 'F': sheet[f'F{row}'].value, 'H': sheet[f'H{row}'].value, 'I': sheet[f'I{row}'].value, 'G': sheet[f'G{row}'].value}
        sheet_data.append(row_data)

    max_cycle = max(cell['F'] for cell in sheet_data if cell['F'] is not None and isinstance(cell['F'], (int, float)))

    input_book = Workbook()
    input_sheet = input_book.active
    input_sheet.title = "Data"
    input_sheet['A1'] = 'CV'
    input_sheet['B1'] = 'SOH'
    input_sheet['C1'] = 'cycleIndex'

    with Pool() as pool:

        # 提交所有的 check_cycle 到进程池
        results = pool.starmap(process_cycle, [(check_cycle, sheet_data, step_index, standard_sapacity) for check_cycle in range(1, max_cycle+1)])
        # 处理返回的结果
        write_row = 2
        for result in results:
            if result:
                check_cycle, final_cv, soh = result
                if soh != 0:
                    input_sheet[f'A{write_row}'] = final_cv
                    input_sheet[f'B{write_row}'] = soh
                    input_sheet[f'C{write_row}'] = check_cycle
                    write_row = write_row + 1


    print("writing!\n")
    input_book.save(output_name)

    # img_window(output_name, step_index)

    print("done!\n")

def browse_file():
    filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    file_path_entry.delete(0, tk.END)
    file_path_entry.insert(0, filename)

def browse_output_file():
    filename = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                            filetypes=[("Excel files", "*.xlsx;*.xls")])
    output_path_entry.delete(0, tk.END)
    output_path_entry.insert(0, filename)

def run_program():
    messagebox.showinfo("提示", "程序执行时间较长，时长受限于cpu性能，点击确定键开始运行")
    work_name = file_path_entry.get()
    output_name = output_path_entry.get()
    step_index = step_index_entry.get()
    standard_sapacity = standard_sapacity_entry.get()
    if work_name:
        print("step_index: "+step_index)
        main(work_name, int(step_index), output_name, float(standard_sapacity))
        messagebox.showinfo("完成", "数据处理完成，结果已保存在"+output_name+" 文件中。")

    else:
        messagebox.showerror("错误", "文件名为空！")

if __name__ == '__main__':
    freeze_support()

    root = tk.Tk()
    root.title("excel数据处理程序")

    tk.Label(root, text="文件路径:").grid(row=0, column=0)
    file_path_entry = tk.Entry(root, width=30)
    file_path_entry.grid(row=0, column=1)
    browse_button = tk.Button(root, text="浏览", command=browse_file)
    browse_button.grid(row=0, column=2)

    tk.Label(root, text="选择步骤:").grid(row=3, column=0)
    step_index_entry = tk.Entry(root, width=20)
    step_index_entry.grid(row=3, column=1)

    tk.Label(root, text="Standard_Capacity:").grid(row=4, column=0)
    standard_sapacity_entry = tk.Entry(root, width=20)
    standard_sapacity_entry.grid(row=4, column=1)

    tk.Label(root, text="文件输出路径:").grid(row=5, column=0)  # Adjust the row index accordingly
    output_path_entry = tk.Entry(root, width=30)
    output_path_entry.grid(row=5, column=1)  # Adjust the row index accordingly
    browse_output_button = tk.Button(root, text="浏览", command=browse_output_file)
    browse_output_button.grid(row=5, column=2)  # Adjust the row index accordingly

    run_button = tk.Button(root, text="启动启动！", command=run_program)
    run_button.grid(row=6, column=1)

    root.mainloop()